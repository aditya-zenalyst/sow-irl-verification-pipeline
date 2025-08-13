"""
Excel file reader with support for multiple formats and encodings
"""

import pandas as pd
import openpyxl
import xlrd
import numpy as np
from pathlib import Path
from typing import Dict, Any, Optional, List, Union
import logging
import chardet
from io import BytesIO
import warnings
warnings.filterwarnings('ignore', category=UserWarning)

from .exceptions import FileReadError


class ExcelReader:
    """Reader for various Excel file formats"""
    
    def __init__(self, config: Any):
        self.config = config
        self.logger = logging.getLogger(__name__)
        
    def read_file(self, file_path: Union[str, Path]) -> Dict[str, Any]:
        """Read Excel file and return raw data"""
        file_path = Path(file_path)
        
        if not file_path.exists():
            raise FileReadError(f"File not found: {file_path}")
            
        file_extension = file_path.suffix.lower()
        
        result = {
            "file_path": str(file_path),
            "file_name": file_path.name,
            "sheets": {},
            "metadata": {
                "file_size": file_path.stat().st_size,
                "file_extension": file_extension
            }
        }
        
        try:
            if file_extension in ['.xlsx', '.xlsm', '.xlsb']:
                result["sheets"] = self.read_xlsx(file_path)
            elif file_extension == '.xls':
                result["sheets"] = self.read_xls(file_path)
            elif file_extension == '.csv':
                result["sheets"] = self.read_csv(file_path)
            else:
                # Try to read as Excel anyway
                result["sheets"] = self.read_with_pandas(file_path)
                
        except Exception as e:
            self.logger.error(f"Primary read failed for {file_path}: {str(e)}")
            # Try fallback methods
            result["sheets"] = self.read_with_fallback(file_path)
            
        return result
    
    def read_xlsx(self, file_path: Path) -> Dict[str, Any]:
        """Read modern Excel files (.xlsx, .xlsm, .xlsb)"""
        sheets = {}
        
        try:
            # Try with openpyxl first (more control)
            wb = openpyxl.load_workbook(file_path, data_only=True, read_only=False)
            
            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                sheet_data = self.extract_openpyxl_data(ws)
                sheets[sheet_name] = sheet_data
                
            wb.close()
            
        except Exception as e:
            self.logger.warning(f"Openpyxl failed, trying pandas: {str(e)}")
            # Fallback to pandas
            sheets = self.read_with_pandas(file_path)
            
        return sheets
    
    def extract_openpyxl_data(self, worksheet) -> Dict[str, Any]:
        """Extract data from openpyxl worksheet"""
        data = []
        max_row = worksheet.max_row
        max_col = worksheet.max_column
        
        # Get all cell values including formulas and merged cells
        for row_idx in range(1, max_row + 1):
            row_data = []
            for col_idx in range(1, max_col + 1):
                cell = worksheet.cell(row=row_idx, column=col_idx)
                value = cell.value
                
                # Handle different cell types
                if value is None:
                    row_data.append(None)
                elif hasattr(value, 'date'):
                    row_data.append(value.isoformat())
                else:
                    row_data.append(value)
                    
            data.append(row_data)
            
        # Get merged cell ranges
        merged_cells = []
        for merged_range in worksheet.merged_cells.ranges:
            merged_cells.append({
                "range": str(merged_range),
                "min_row": merged_range.min_row,
                "max_row": merged_range.max_row,
                "min_col": merged_range.min_col,
                "max_col": merged_range.max_col
            })
            
        return {
            "data": data,
            "shape": (max_row, max_col),
            "merged_cells": merged_cells,
            "has_formulas": any(
                worksheet.cell(row=r, column=c).data_type == 'f'
                for r in range(1, min(10, max_row + 1))
                for c in range(1, min(10, max_col + 1))
            )
        }
    
    def read_xls(self, file_path: Path) -> Dict[str, Any]:
        """Read old Excel files (.xls)"""
        sheets = {}
        
        try:
            # Use xlrd for old Excel format
            workbook = xlrd.open_workbook(file_path, on_demand=True)
            
            for sheet_name in workbook.sheet_names():
                sheet = workbook.sheet_by_name(sheet_name)
                sheet_data = self.extract_xlrd_data(sheet)
                sheets[sheet_name] = sheet_data
                
            workbook.release_resources()
            
        except Exception as e:
            self.logger.warning(f"xlrd failed, trying pandas: {str(e)}")
            sheets = self.read_with_pandas(file_path)
            
        return sheets
    
    def extract_xlrd_data(self, sheet) -> Dict[str, Any]:
        """Extract data from xlrd sheet"""
        data = []
        
        for row_idx in range(sheet.nrows):
            row_data = []
            for col_idx in range(sheet.ncols):
                cell = sheet.cell(row_idx, col_idx)
                
                # Convert cell types
                if cell.ctype == xlrd.XL_CELL_EMPTY:
                    value = None
                elif cell.ctype == xlrd.XL_CELL_DATE:
                    value = xlrd.xldate_as_datetime(cell.value, 0).isoformat()
                elif cell.ctype == xlrd.XL_CELL_BOOLEAN:
                    value = bool(cell.value)
                elif cell.ctype == xlrd.XL_CELL_ERROR:
                    value = f"ERROR: {cell.value}"
                else:
                    value = cell.value
                    
                row_data.append(value)
                
            data.append(row_data)
            
        return {
            "data": data,
            "shape": (sheet.nrows, sheet.ncols),
            "merged_cells": []
        }
    
    def read_csv(self, file_path: Path) -> Dict[str, Any]:
        """Read CSV files"""
        sheets = {}
        
        # Detect encoding
        with open(file_path, 'rb') as f:
            raw_data = f.read(10000)
            detected = chardet.detect(raw_data)
            encoding = detected.get('encoding', 'utf-8')
            
        try:
            # Try different delimiters
            for delimiter in [',', ';', '\t', '|']:
                try:
                    df = pd.read_csv(
                        file_path, 
                        encoding=encoding,
                        delimiter=delimiter,
                        engine='python',
                        on_bad_lines='skip'
                    )
                    
                    if len(df.columns) > 1:  # Valid delimiter found
                        break
                except:
                    continue
                    
            # Convert to list format
            data = df.values.tolist()
            headers = df.columns.tolist()
            data.insert(0, headers)
            
            sheets["Sheet1"] = {
                "data": data,
                "shape": (len(data), len(headers)),
                "merged_cells": []
            }
            
        except Exception as e:
            self.logger.error(f"CSV read failed: {str(e)}")
            # Return raw text as single column
            with open(file_path, 'r', encoding=encoding, errors='ignore') as f:
                lines = f.readlines()
                data = [[line.strip()] for line in lines]
                
            sheets["Sheet1"] = {
                "data": data,
                "shape": (len(data), 1),
                "merged_cells": []
            }
            
        return sheets
    
    def read_with_pandas(self, file_path: Path) -> Dict[str, Any]:
        """Read Excel file using pandas"""
        sheets = {}
        
        try:
            # Read all sheets
            excel_file = pd.ExcelFile(file_path, engine=None)
            
            for sheet_name in excel_file.sheet_names:
                df = pd.read_excel(
                    excel_file, 
                    sheet_name=sheet_name,
                    header=None,  # Don't assume headers
                    dtype=object,  # Keep all data as objects initially
                    na_filter=True,
                    keep_default_na=True
                )
                
                # Convert to list format
                data = df.replace({np.nan: None}).values.tolist()
                
                sheets[sheet_name] = {
                    "data": data,
                    "shape": df.shape,
                    "merged_cells": []
                }
                
        except Exception as e:
            self.logger.error(f"Pandas read failed: {str(e)}")
            raise FileReadError(f"Unable to read file: {str(e)}")
            
        return sheets
    
    def read_with_fallback(self, file_path: Path) -> Dict[str, Any]:
        """Try multiple methods to read the file"""
        methods = [
            ("pandas_excel", lambda: self.read_with_pandas(file_path)),
            ("openpyxl", lambda: self.read_xlsx(file_path)),
            ("xlrd", lambda: self.read_xls(file_path)),
            ("csv", lambda: self.read_csv(file_path))
        ]
        
        for method_name, method_func in methods:
            try:
                self.logger.info(f"Trying fallback method: {method_name}")
                return method_func()
            except Exception as e:
                self.logger.warning(f"Fallback {method_name} failed: {str(e)}")
                continue
                
        raise FileReadError(f"All read methods failed for {file_path}")
    
    def read_sheet_range(self, file_path: Path, sheet_name: str, 
                        range_str: str) -> List[List[Any]]:
        """Read specific range from a sheet"""
        try:
            df = pd.read_excel(
                file_path,
                sheet_name=sheet_name,
                usecols=range_str,
                header=None
            )
            return df.values.tolist()
        except:
            return []