#!/usr/bin/env python3
"""
IRL DD Pipeline - Information Requirements List Generator
========================================================
Takes SOW LLM output and generates specific, actionable data requests
in Excel format following institutional standards.

Features:
- Reads SOW LLM output files automatically
- Extracts company name and financial periods
- Generates structured IRL in exact sample format
- Creates Excel output with version tracking
- Maps DD procedures to specific data requests
"""

import os
import sys
import json
import logging
from typing import Dict, Any, List, Optional
from pathlib import Path
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from datetime import datetime

# Configure logging
logging.basicConfig(level=logging.INFO, format='%(levelname)s:%(name)s:%(message)s')
logger = logging.getLogger(__name__)

class IRLDueDiligencePipeline:
    """
    Complete pipeline for generating Information Requirements List from SOW output
    """
    
    def __init__(self):
        self.version_file = "irl_version_tracker.json"
        self.version_tracker = self._load_version_tracker()
        # Copy API configuration from SOW LLM
        self.api_key = self._load_api_key()
    
    def _load_version_tracker(self) -> Dict[str, int]:
        """Load IRL version tracker from file"""
        if os.path.exists(self.version_file):
            try:
                with open(self.version_file, "r") as f:
                    return json.load(f)
            except:
                return {}
        return {}
    
    def _save_version_tracker(self):
        """Save IRL version tracker to file"""
        try:
            with open(self.version_file, "w") as f:
                json.dump(self.version_tracker, f, indent=2)
        except Exception as e:
            logger.warning(f"Could not save IRL version tracker: {e}")
    
    def _load_api_key(self) -> Optional[str]:
        """Load API key from SOW LLM .env file with multiple path attempts"""
        # Try multiple possible paths for the .env file
        possible_paths = [
            "../SOW LLM/.env",
            "/Users/adarshsingh/Desktop/NEW LLM ZEN/SOW LLM/.env",
            ".env",
            "SOW LLM/.env"
        ]
        
        for sow_env_path in possible_paths:
            if os.path.exists(sow_env_path):
                try:
                    with open(sow_env_path, 'r') as f:
                        for line in f:
                            if line.startswith('CLAUDE_API_KEY='):
                                api_key = line.split('=', 1)[1].strip()
                                if api_key:
                                    logger.info(f"✅ API key loaded from: {sow_env_path}")
                                    return api_key
                except Exception as e:
                    logger.warning(f"Failed to read {sow_env_path}: {e}")
                    continue
        
        # Also try environment variable
        env_api_key = os.getenv('CLAUDE_API_KEY')
        if env_api_key:
            logger.info("✅ API key loaded from environment variable")
            return env_api_key
        
        logger.error("❌ No API key found in any location")
        return None
    
    def read_sow_output(self, sow_file_path: str) -> Dict[str, Any]:
        """Read and parse SOW LLM output file"""
        logger.info(f"Reading SOW output: {sow_file_path}")
        
        try:
            with open(sow_file_path, 'r', encoding='utf-8') as f:
                sow_content = f.read()
            
            # Extract company name from SOW file
            company_name = self._extract_company_name_from_sow(sow_content)
            
            # Extract financial periods
            financial_periods = self._extract_financial_periods(sow_content)
            
            # Extract DD sections
            dd_sections = self._extract_dd_sections(sow_content)
            
            return {
                "status": "success",
                "company_name": company_name,
                "financial_periods": financial_periods,
                "dd_sections": dd_sections,
                "full_content": sow_content
            }
            
        except Exception as e:
            logger.error(f"Error reading SOW file: {e}")
            return {"status": "error", "error": str(e)}
    
    def _extract_company_name_from_sow(self, sow_content: str) -> str:
        """Extract company name from SOW content with comprehensive pattern matching"""
        import re
        
        # Step 1: Look for "Company: NAME" pattern first (most reliable)
        company_line_match = re.search(r'Company:\s*([^\n\r]+)', sow_content, re.IGNORECASE)
        if company_line_match:
            company_line = company_line_match.group(1).strip()
            # Take only the first word/token for simple names like "ABC"
            first_token = company_line.split()[0] if company_line.split() else company_line
            if self._is_valid_company_name(first_token):
                return first_token
            # If first token fails, try the cleaned full line
            cleaned_name = self._clean_extracted_company_name(company_line, sow_content)
            if self._is_valid_company_name(cleaned_name):
                return cleaned_name
        
        # Step 2: Try various extraction patterns from SOW content  
        extraction_patterns = [
            # Pattern 1: Analysis/statements for company
            r'(?:statements provided for|analysis of|analysis for)\s+([A-Z][a-zA-Z\s&]+(?:Private Limited|Pvt\.?\s*Ltd\.?|Limited|Corporation))',
            
            # Pattern 2: Direct company name patterns
            r'([A-Z][a-zA-Z\s&]+(?:Private Limited|Pvt\.?\s*Ltd\.?|Limited|Corporation|LLC|Inc\.?))',
            
            # Pattern 3: Quoted company names
            r'"([A-Z][A-Za-z\s&]+(?:Private Limited|Pvt\.?\s*Ltd\.?|Limited|Corporation|LLC|Inc\.?))"',
            
            # Pattern 4: Company prefix patterns  
            r'Company[:\s]+"?([A-Z][a-zA-Z\s&]+(?:Private Limited|Pvt\.?\s*Ltd\.?|Limited|Corporation|LLC|Inc\.?))"?',
            
            # Pattern 5: Short names in quotes
            r'"([A-Z]{2,10})"',
            
            # Pattern 6: CIN pattern (Indian companies)
            r'([A-Z][a-zA-Z\s&]+(?:Private Limited|Pvt\.?\s*Ltd\.?)).*?CIN[:\s]*[A-Z0-9]+'
        ]
        
        for pattern in extraction_patterns:
            matches = re.findall(pattern, sow_content, re.IGNORECASE)
            for match in matches:
                match = match.strip()
                cleaned_name = self._clean_extracted_company_name(match, sow_content)
                if self._is_valid_company_name(cleaned_name):
                    return cleaned_name
        
        # Step 3: Fallback - look for reasonable company names in first few lines
        lines = sow_content.split('\n')
        for line in lines[:10]:
            line = line.strip()
            if line and len(line) > 2 and len(line) < 100:
                if self._is_valid_company_name(line):
                    return line
        
        return "Target Company"
    
    def _clean_extracted_company_name(self, raw_name: str, full_content: str) -> str:
        """Clean and validate extracted company name"""
        import re
        
        company_name = raw_name.strip()
        
        # Remove common problematic prefixes
        prefixes_to_remove = [
            "Company Name: ", "COMPANY NAME: ", "Company: ", "Company name: ",
            "Based on the provided document", "Since only the company name ",
            "The company name is: ", "Analysis of ", "Statements provided for "
        ]
        
        for prefix in prefixes_to_remove:
            if company_name.startswith(prefix):
                company_name = company_name[len(prefix):].strip()
                break
        
        # Handle quoted names with descriptive text
        # "Since only the company name "ABC" is provided without..."
        quoted_match = re.search(r'"([^"]+)"', company_name)
        if quoted_match:
            potential_name = quoted_match.group(1).strip()
            if self._is_valid_company_name(potential_name):
                return potential_name
        
        # Remove descriptive trailing text
        cleanup_patterns = [
            r'\s+is provided without.*$',
            r'\s+demonstrates.*$',
            r'\s+shows.*$',
            r'\s+registered in.*$',
            r'\s*\(CIN:.*\).*$'
        ]
        
        for pattern in cleanup_patterns:
            company_name = re.sub(pattern, '', company_name, flags=re.IGNORECASE).strip()
        
        # Remove quotes and normalize
        company_name = company_name.replace('"', '').replace("'", "")
        company_name = re.sub(r'\s+', ' ', company_name).strip()
        
        # Handle comma-separated explanations
        if "," in company_name and len(company_name) > 50:
            parts = company_name.split(",")
            for part in parts:
                part = part.strip()
                if self._is_valid_company_name(part):
                    return part
        
        return company_name
    
    def _is_valid_company_name(self, name: str) -> bool:
        """Validate if a string is a proper company name"""
        if not name or len(name.strip()) < 2:
            return False
            
        name = name.strip()
        name_lower = name.lower()
        
        # Exclude invalid phrases
        invalid_phrases = {
            'based on', 'provided', 'document', 'financial', 'statements',
            'analysis', 'comprehensive', 'the', 'and', 'or', 'is', 'are',
            'with', 'without', 'complete', 'since', 'only', 'target company'
        }
        
        # Check if it's just an invalid phrase
        if name_lower in invalid_phrases:
            return False
        
        # Check for multiple invalid phrases
        invalid_count = sum(1 for phrase in invalid_phrases if phrase in name_lower)
        if invalid_count > 1:
            return False
        
        # Positive company indicators
        company_indicators = [
            'private limited', 'ltd', 'limited', 'corporation', 'corp',
            'llc', 'inc', 'pvt', 'solutions', 'systems', 'technologies',
            ' & ', 'group', 'holdings', 'enterprises'
        ]
        
        # Strong positive signal
        if any(indicator in name_lower for indicator in company_indicators):
            return True
        
        # All caps short names (like "ABC")
        if name.isupper() and 2 <= len(name) <= 10 and name.isalpha():
            return True
        
        # Reasonable length with proper capitalization
        if 3 <= len(name) <= 100 and name[0].isupper():
            return True
        
        return False
    
    def _extract_financial_periods(self, sow_content: str) -> Dict[str, str]:
        """Extract financial periods mentioned in SOW"""
        periods = {"current_year": "", "previous_year": "", "balance_sheet_date": ""}
        
        # Look for period patterns in SOW
        import re
        
        # Look for FY patterns
        fy_matches = re.findall(r'FY\s*(\d{4}[-/]\d{2,4})', sow_content)
        if fy_matches:
            periods["current_year"] = fy_matches[-1] if fy_matches else ""
            periods["previous_year"] = fy_matches[-2] if len(fy_matches) > 1 else ""
        
        # Look for March dates (common in Indian companies)
        march_matches = re.findall(r'March\s+\d{1,2},?\s+(\d{4})', sow_content)
        if march_matches:
            periods["balance_sheet_date"] = f"March 31, {march_matches[-1]}"
        
        # Look for year ranges
        year_matches = re.findall(r'(\d{4}[-/]\d{4})', sow_content)
        if year_matches and not periods["current_year"]:
            periods["current_year"] = year_matches[-1]
        
        return periods
    
    def _extract_dd_sections(self, sow_content: str) -> List[Dict[str, Any]]:
        """Extract DD sections and procedures from SOW"""
        sections = []
        
        # Look for table format sections
        lines = sow_content.split('\n')
        current_section = None
        in_table = False
        
        for line in lines:
            line = line.strip()
            
            # Detect table start
            if '| Analysis Area | Detailed Procedures |' in line:
                in_table = True
                continue
            
            # Detect table separator
            if in_table and line.startswith('|') and '---' in line:
                continue
            
            # Extract table rows
            if in_table and line.startswith('| **') and '**' in line:
                parts = line.split('|')
                if len(parts) >= 3:
                    section_name = parts[1].replace('**', '').strip()
                    procedures = parts[2].strip()
                    
                    sections.append({
                        "name": section_name,
                        "procedures": procedures,
                        "priority": self._assign_section_priority(section_name)
                    })
        
        return sections
    
    def _assign_section_priority(self, section_name: str) -> str:
        """Assign priority based on section importance"""
        high_priority = [
            "Quality of Earnings Analysis",
            "Income Statement Analysis", 
            "Working Capital Management",
            "Cash Flow Analysis",
            "Balance Sheet Review"
        ]
        
        medium_priority = [
            "Capital Structure & Debt Analysis",
            "General Overview & Financial Reporting",
            "Accounting Policies & Estimates"
        ]
        
        if any(hp in section_name for hp in high_priority):
            return "High"
        elif any(mp in section_name for mp in medium_priority):
            return "Medium"
        else:
            return "Low"
    
    def generate_irl_from_sow(self, sow_data: Dict[str, Any]) -> Dict[str, Any]:
        """Generate IRL content from parsed SOW data"""
        logger.info("Generating IRL from SOW data...")
        
        try:
            # Generate IRL using LLM
            irl_content = self._generate_irl_with_llm(sow_data)
            
            if not irl_content:
                return {"status": "error", "error": "Failed to generate IRL content"}
            
            # Structure IRL data
            structured_irl = self._structure_irl_data(irl_content, sow_data)
            
            return {
                "status": "success",
                "irl_data": structured_irl,
                "company_name": sow_data["company_name"],
                "financial_periods": sow_data["financial_periods"]
            }
            
        except Exception as e:
            logger.error(f"Error generating IRL: {e}")
            return {"status": "error", "error": str(e)}
    
    def _generate_irl_with_llm(self, sow_data: Dict[str, Any]) -> str:
        """Generate dynamic IRL based on actual SOW content and scope"""
        logger.info("🔄 Starting dynamic IRL generation based on SOW scope...")
        
        company_name = sow_data['company_name']
        financial_periods = sow_data['financial_periods']
        dd_sections = sow_data['dd_sections']
        full_sow_content = sow_data['full_content']
        
        # Analyze SOW to determine actual scope
        scope_analysis = self._analyze_sow_scope(full_sow_content, dd_sections)
        logger.info(f"📊 Scope analysis: {scope_analysis['total_areas']} areas, {scope_analysis['estimated_requests']} estimated requests")
        
        # Choose generation approach based on scope size
        if len(dd_sections) > 8:
            logger.info("📋 Using multi-pass approach for comprehensive DD scope...")
            dynamic_irl = self._generate_comprehensive_irl_multipass(
                company_name, financial_periods, dd_sections, full_sow_content, scope_analysis
            )
        else:
            logger.info("📋 Generating focused IRL based on SOW scope...")
            dynamic_irl = self._generate_dynamic_irl_from_sow(
                company_name, financial_periods, dd_sections, full_sow_content, scope_analysis
            )
        
        logger.info(f"✅ Dynamic IRL generation complete: {len(dynamic_irl)} characters")
        return dynamic_irl
    
    def _format_dd_sections_for_prompt(self, dd_sections: List[Dict[str, Any]]) -> str:
        """Format DD sections for LLM prompt"""
        formatted = ""
        for section in dd_sections:
            formatted += f"\n**{section['name']}**:\n{section['procedures']}\n"
        return formatted
    
    def _calculate_required_tokens(self, dd_sections: List[Dict[str, Any]], scope_analysis: Dict[str, Any]) -> int:
        """Calculate required tokens based on DD scope size"""
        num_areas = len(dd_sections)
        
        if num_areas <= 3:
            # Focused scope (like revenue-only)
            return 4000
        elif num_areas <= 8:
            # Medium scope 
            return 6000
        elif num_areas <= 12:
            # Large scope
            return 7000
        else:
            # Comprehensive scope (14+ areas) - max Claude limit
            return 8000  # Maximum safe limit for Claude
    
    def _direct_llm_call(self, prompt: str, max_tokens: int = 4000) -> Dict[str, Any]:
        """Direct LLM API call for IRL generation"""
        import requests
        
        if not self.api_key:
            return {"status": "error", "error": "API key not found"}
        
        try:
            headers = {
                "x-api-key": self.api_key,
                "content-type": "application/json",
                "anthropic-version": "2023-06-01"
            }
            
            payload = {
                "model": "claude-3-5-sonnet-20241022",
                "max_tokens": max_tokens,
                "messages": [{"role": "user", "content": prompt}]
            }
            
            response = requests.post(
                "https://api.anthropic.com/v1/messages",
                headers=headers,
                json=payload,
                timeout=60
            )
            
            if response.status_code == 200:
                result = response.json()
                return {
                    "status": "success",
                    "analysis": result['content'][0]['text']
                }
            else:
                return {"status": "error", "error": f"API error: {response.status_code}"}
                
        except Exception as e:
            return {"status": "error", "error": str(e)}
    
    def _structure_irl_data(self, irl_content: str, sow_data: Dict[str, Any]) -> List[Dict[str, Any]]:
        """Structure LLM-generated IRL content into Excel format"""
        structured_data = []
        
        # Add header rows with financial periods
        periods = sow_data["financial_periods"]
        current_year = periods.get("current_year", "FY2015")
        previous_year = periods.get("previous_year", "FY2014") 
        balance_date = periods.get("balance_sheet_date", "March 31, 2015")
        
        # FIXED: Don't add header information as regular data items
        # These will be handled directly in the Excel header section
        
        # Parse LLM-generated IRL content and inject section headers
        if irl_content:
            parsed_requests = self._parse_irl_content_with_headers(irl_content, sow_data["dd_sections"])
            structured_data.extend(parsed_requests)
        else:
            # Fallback: Generate basic requests from DD sections
            structured_data.extend(self._generate_basic_irl_from_dd_sections(sow_data["dd_sections"]))
        
        return structured_data
    
    def _parse_irl_content_with_headers(self, irl_content: str, dd_sections: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
        """Parse IRL content and intelligently inject proper IRL section headers"""
        # First parse the content normally
        parsed_requests = self._parse_irl_content(irl_content)
        
        if not parsed_requests or not dd_sections:
            return parsed_requests
        
        # Calculate requests per section (distribute evenly)
        total_requests = len(parsed_requests)
        num_sections = len(dd_sections)
        requests_per_section = max(1, total_requests // num_sections)
        
        # Inject section headers at appropriate intervals
        structured_requests = []
        request_count = 0
        section_index = 0
        
        for i, request in enumerate(parsed_requests):
            # Add section header at the beginning of each section
            if request_count == 0 or (request_count % requests_per_section == 0 and section_index < num_sections):
                if section_index < num_sections:
                    sow_section_name = dd_sections[section_index].get('name', f'DD Area {section_index + 1}')
                    # Map SOW section to proper IRL section header  
                    irl_section_name = self._map_sow_to_irl_section(sow_section_name)
                    
                    # Add section header
                    structured_requests.append({
                        "id": "",
                        "info_request": f"**{irl_section_name}**",
                        "priority": "",
                        "status": "",
                        "zenalyst_remarks": "",
                        "management_remarks": ""
                    })
                    section_index += 1
            
            # Add the actual request
            structured_requests.append(request)
            request_count += 1
        
        return structured_requests
    
    def _map_sow_to_irl_section(self, sow_section_name: str) -> str:
        """Map SOW DD section names to standard IRL section headers"""
        sow_name_lower = sow_section_name.lower()
        
        # Standard IRL section mappings
        if any(keyword in sow_name_lower for keyword in ['quality of earnings', 'revenue', 'earnings']):
            return "REVENUE ANALYSIS"
        elif any(keyword in sow_name_lower for keyword in ['income statement', 'profitability', 'margin']):
            return "INCOME STATEMENT ANALYSIS"
        elif any(keyword in sow_name_lower for keyword in ['working capital', 'liquidity']):
            return "WORKING CAPITAL MANAGEMENT"
        elif any(keyword in sow_name_lower for keyword in ['cash flow', 'cash']):
            return "CASH FLOW ANALYSIS"
        elif any(keyword in sow_name_lower for keyword in ['balance sheet', 'asset', 'position']):
            return "BALANCE SHEET REVIEW"
        elif any(keyword in sow_name_lower for keyword in ['capital structure', 'debt', 'financing']):
            return "CAPITAL STRUCTURE & DEBT ANALYSIS"
        elif any(keyword in sow_name_lower for keyword in ['general', 'reporting', 'systems']):
            return "GENERAL OVERVIEW & FINANCIAL REPORTING"
        elif any(keyword in sow_name_lower for keyword in ['accounting', 'policies', 'estimates']):
            return "ACCOUNTING POLICIES & ESTIMATES"
        elif any(keyword in sow_name_lower for keyword in ['compensation', 'payroll', 'benefits']):
            return "COMPENSATION, PAYROLL & BENEFITS"
        elif any(keyword in sow_name_lower for keyword in ['related party', 'transactions']):
            return "RELATED PARTY TRANSACTIONS"
        elif any(keyword in sow_name_lower for keyword in ['tax', 'taxation']):
            return "TAX MATTERS"
        elif any(keyword in sow_name_lower for keyword in ['contingent', 'liabilities', 'commitments']):
            return "CONTINGENT LIABILITIES & COMMITMENTS"
        elif any(keyword in sow_name_lower for keyword in ['operational', 'operations']):
            return "OPERATIONAL ANALYSIS"
        elif any(keyword in sow_name_lower for keyword in ['risk', 'assessment']):
            return "RISK ASSESSMENT"
        elif any(keyword in sow_name_lower for keyword in ['customer', 'market']):
            return "CUSTOMER ANALYSIS"
        else:
            # Fallback: clean up the original name
            clean_name = sow_section_name.replace('(Priority:', '').replace('High)', '').replace('Medium)', '').replace('Low)', '').strip()
            return clean_name.upper()
    
    def _parse_irl_content(self, irl_content: str) -> List[Dict[str, Any]]:
        """Parse LLM-generated IRL content preserving full structure with sub-points"""
        requests = []
        lines = irl_content.split('\n')
        
        current_request = None
        current_request_text = []
        
        for line in lines:
            line = line.strip()
            
            # Skip empty lines, generator instructions, and section headers (we'll add our own)
            if not line or 'Generate' in line or 'Continue' in line:
                continue
            
            # SKIP any section headers - we'll inject our own consistent ones
            if line.startswith('**') and line.endswith('**') and len(line) > 4:
                # Save previous request if exists
                if current_request and current_request_text:
                    final_text = '\n'.join(current_request_text)
                    final_text, priority = self._extract_and_clean_priority(final_text, current_request.get("priority", ""))
                    current_request["info_request"] = final_text
                    current_request["priority"] = priority
                    requests.append(current_request)
                    current_request = None
                    current_request_text = []
                
                # SKIP the header - we'll add consistent ones later
                continue
            
            # Detect numbered requests
            import re
            number_match = re.match(r'^(\d+)\.\s*(.+)', line)
            if number_match:
                # Save previous request if exists
                if current_request and current_request_text:
                    # Clean up the final request text and extract priority from the last line
                    final_text = '\n'.join(current_request_text)
                    final_text, priority = self._extract_and_clean_priority(final_text, current_request.get("priority", ""))
                    current_request["info_request"] = final_text
                    current_request["priority"] = priority
                    requests.append(current_request)
                
                # Start new request
                request_id = number_match.group(1)
                first_line = number_match.group(2)
                
                # Don't extract priority from first line - wait for (f) sub-point
                current_request = {
                    "id": request_id,
                    "info_request": "",
                    "priority": "",
                    "status": "",
                    "zenalyst_remarks": "",
                    "management_remarks": ""
                }
                current_request_text = [first_line] if first_line else []
                
            # Detect sub-points (a), (b), (c), etc.
            elif re.match(r'^\([a-f]\)\s+.+', line) and current_request:
                current_request_text.append(line)
            
            # Handle continuation lines for sub-points
            elif current_request and line and not line.startswith('(') and len(current_request_text) > 0:
                # Check if it's a continuation of the last sub-point
                if not re.match(r'^\d+\.', line) and not line.startswith('**'):  # Not a new numbered item or section header
                    current_request_text.append(line)
        
        # Don't forget the last request
        if current_request and current_request_text:
            final_text = '\n'.join(current_request_text)
            final_text, priority = self._extract_and_clean_priority(final_text, current_request.get("priority", ""))
            current_request["info_request"] = final_text
            current_request["priority"] = priority
            requests.append(current_request)
        
        return requests
    
    def _extract_and_clean_priority(self, text: str, existing_priority: str) -> tuple:
        """Extract priority from text and clean up duplicates"""
        import re
        
        # Find all priority mentions
        priority_matches = re.findall(r'\(Priority:\s*(High|Medium|Low)\)', text)
        
        # Remove all priority mentions from text
        cleaned_text = re.sub(r'\s*\(Priority:\s*(High|Medium|Low)\)', '', text)
        
        # Determine final priority
        if priority_matches:
            # Use the last priority mentioned (usually from (f) sub-point)
            final_priority = priority_matches[-1]
        elif existing_priority:
            final_priority = existing_priority
        else:
            # Assign based on content
            final_priority = self._assign_request_priority(cleaned_text)
        
        return cleaned_text.strip(), final_priority
    
    def _generate_section_a(self, company_name: str, financial_periods: Dict[str, str], dd_sections: List[Dict[str, Any]]) -> str:
        """Generate Section A: Financial statements, MIS and general information (requests 1-15)"""
        
        # Extract relevant DD sections for Section A
        section_a_keywords = ["financial statements", "earnings", "income statement", "revenue", "general"]
        relevant_sections = [s for s in dd_sections if any(keyword in s.get("name", "").lower() for keyword in section_a_keywords)]
        
        # Format periods for prompt
        periods_text = f"FY{financial_periods.get('current_year', '2014-2015')}, FY{financial_periods.get('previous_year', '2013-2014')}"
        
        prompt = f"""GENERATE SECTION A ONLY - FINANCIAL STATEMENTS AND GENERAL INFORMATION

**SECTION A: Financial statements, MIS and other general information**

Company: {company_name}
Historical periods: {periods_text}

RELEVANT DD FOCUS AREAS:
{self._format_dd_sections_for_prompt(relevant_sections)}

Generate requests 1-15 ONLY with detailed sub-points (a), (b), (c), (d), (e), (f).

MANDATORY FORMAT:
1. (a) Excel copies of standalone and consolidated financial statements (profit and loss statement, balance sheet and cash flow statements), linked to the detailed trial balances for the historical period.
(b) Copies of standalone and consolidated audited financial statements along with audit report for the historical period with complete notes to accounts.
(c) Board resolutions approving the financial statements along with minutes of audit committee meetings discussing financial results.
(d) Reconciliation between provisional financial statements and final audited statements with explanations for all adjustments.
(e) Independent auditor's management letter, internal control observations, and any qualifications or emphasis of matter.
(f) Statutory compliance certificates and regulatory filings covering all applicable laws.

2. (a) Monthly or quarterly management accounts (MIS) prepared by the management including detailed variance analysis, KPIs as tracked by the Management, operational statistics for the historical period in Excel format.
(b) Budget versus actual analysis with detailed variance explanations for all material deviations exceeding 5% or significant absolute amounts.
(c) Flash reports, CEO dashboards, and management presentations used for internal decision making during the historical period.
(d) Key performance indicators (KPIs) as tracked by management including operational metrics, efficiency ratios, and business-specific parameters.
(e) Reconciliation between the MIS and consolidated financial statements with explanations for all differences.
(f) Management reporting packages and forecast models used for internal performance tracking.

Continue this EXACT format for requests 3-15. Tailor content to {company_name}'s specific business and the DD focus areas above.

GENERATE ALL 15 REQUESTS NOW - DO NOT STOP OR ASK FOR CONFIRMATION."""

        result = self._direct_llm_call(prompt)
        return result.get("analysis", "") if result["status"] == "success" else ""
    
    def _format_dd_sections_for_scope_prompt(self, dd_sections: List[Dict[str, Any]]) -> str:
        """Format parsed DD sections for scope-aware IRL generation"""
        if not dd_sections:
            return "Standard comprehensive due diligence procedures"
        
        formatted_sections = []
        for i, section in enumerate(dd_sections, 1):
            section_name = section.get('name', f'DD Area {i}')
            procedures = section.get('procedures', 'Standard procedures')
            
            # Clean up procedures text
            if procedures and len(procedures) > 10:
                # Split procedures into bullet points if they're in numbered format
                procedure_lines = []
                for line in procedures.split('\n'):
                    line = line.strip()
                    if line and (line[0].isdigit() or line.startswith('-') or line.startswith('•')):
                        procedure_lines.append(f"  - {line}")
                    elif line:
                        procedure_lines.append(f"  - {line}")
                
                formatted_procedures = '\n'.join(procedure_lines[:8])  # Limit per section
            else:
                formatted_procedures = f"  - {procedures}"
            
            formatted_sections.append(f"**{section_name}**:\n{formatted_procedures}")
        
        return '\n\n'.join(formatted_sections)
    
    def _create_scope_specific_instruction(self, dd_sections: List[Dict[str, Any]], scope_analysis: Dict[str, Any]) -> str:
        """Create specific instructions based on the DD scope type"""
        if not dd_sections:
            return "Generate standard comprehensive due diligence requirements."
        
        section_names = [section.get('name', '').lower() for section in dd_sections]
        
        # Revenue-focused scope
        if any('revenue' in name or 'earnings' in name or 'customer' in name for name in section_names):
            return """- Focus EXCLUSIVELY on revenue quality, customer analysis, and earnings validation
- DO NOT include balance sheet, cash flow, or operational DD areas not specified
- Emphasize customer contracts, revenue recognition, and sales analytics
- Generate 4-5 requests per revenue-related DD area only"""
        
        # Working capital focused
        elif any('working capital' in name or 'cash flow' in name for name in section_names):
            return """- Focus EXCLUSIVELY on working capital components and cash flow analysis
- DO NOT include revenue, HR, or strategic DD areas not specified
- Emphasize receivables, payables, inventory, and cash management
- Generate 4-5 requests per working capital DD area only"""
        
        # Balance sheet focused
        elif any('balance sheet' in name or 'assets' in name or 'liabilities' in name for name in section_names):
            return """- Focus EXCLUSIVELY on balance sheet components and asset/liability analysis
- DO NOT include P&L, operations, or strategic DD areas not specified
- Emphasize asset verification, liability validation, and balance sheet reconciliation
- Generate 4-5 requests per balance sheet DD area only"""
        
        # Comprehensive scope
        elif len(dd_sections) > 8:
            return """- This is a comprehensive DD scope covering multiple areas
- Generate 3-5 requests per DD area mentioned
- Ensure coverage of all specified DD areas without adding unlisted areas"""
        
        # Focused scope
        else:
            return f"""- This is a focused DD scope with {len(dd_sections)} specific areas
- Generate 4-6 requests per DD area mentioned
- DO NOT add DD areas not explicitly listed in the scope
- Focus exclusively on the {len(dd_sections)} areas specified"""
    
    def _generate_comprehensive_irl_multipass(self, company_name: str, financial_periods: Dict[str, str], 
                                            dd_sections: List[Dict[str, Any]], full_sow_content: str,
                                            scope_analysis: Dict[str, Any]) -> str:
        """Generate comprehensive IRL using multi-pass approach for large DD scopes"""
        
        # Split DD sections into manageable chunks (4-5 areas per pass)
        chunk_size = 4
        dd_chunks = [dd_sections[i:i+chunk_size] for i in range(0, len(dd_sections), chunk_size)]
        
        logger.info(f"🗓️ Splitting {len(dd_sections)} DD areas into {len(dd_chunks)} passes")
        
        all_irl_content = []
        request_counter = 1
        
        for i, chunk in enumerate(dd_chunks, 1):
            logger.info(f"📋 Pass {i}/{len(dd_chunks)}: Generating requests for {len(chunk)} DD areas...")
            
            chunk_content = self._generate_irl_chunk(
                company_name, financial_periods, chunk, scope_analysis, request_counter
            )
            
            if chunk_content:
                all_irl_content.append(chunk_content)
                # Update request counter based on generated content
                request_counter += self._count_requests_in_content(chunk_content)
            
        combined_irl = '\n\n'.join(all_irl_content)
        logger.info(f"✅ Multi-pass comprehensive IRL complete: {len(combined_irl)} characters")
        
        return combined_irl
    
    def _generate_irl_chunk(self, company_name: str, financial_periods: Dict[str, str], 
                           dd_chunk: List[Dict[str, Any]], scope_analysis: Dict[str, Any], 
                           start_request_num: int) -> str:
        """Generate IRL content for a chunk of DD areas"""
        
        # Format periods and DD chunk
        current_year = financial_periods.get('current_year', '2024-2025')
        previous_year = financial_periods.get('previous_year', '2023-2024')
        periods_text = f"FY{current_year}, FY{previous_year}"
        
        formatted_dd_chunk = self._format_dd_sections_for_scope_prompt(dd_chunk)
        
        # Create clean section headers for this chunk
        section_headers = self._create_section_headers_for_chunk(dd_chunk)
        
        prompt = f"""GENERATE IRL SECTION - PART OF COMPREHENSIVE DD

COMPANY: {company_name}
HISTORICAL PERIODS: {periods_text}
START REQUEST NUMBER: {start_request_num}

DD AREAS FOR THIS SECTION:
{formatted_dd_chunk}

MANDATORY SECTION HEADERS (use exactly as shown):
{section_headers}

CRITICAL FORMATTING RULES:
1. Start EACH DD area with its clean section header (NO priority in header)
2. Generate 3-4 information requests per DD area
3. Number requests sequentially starting from {start_request_num}
4. Each request has detailed sub-points (a), (b), (c), (d), (e), (f)
5. Priority goes ONLY at end of (f) sub-point: "(Priority: High)"
6. DO NOT repeat priority - use only once per request
7. Continue numbering across all DD areas (no restart)
8. Tailor all content to {company_name}

STRICT FORMAT:
**[CLEAN DD AREA NAME]**

{start_request_num}. (a) [Specific requirement]
(b) [Supporting documentation]
(c) [Analysis methodology]
(d) [Validation procedures]
(e) [Format specifications]
(f) [Priority: High (with no mention of deadline or Timeline requirements)]

{start_request_num + 1}. (a) [Next requirement for same area]
(b) [Supporting documentation]
(c) [Analysis methodology]
(d) [Validation procedures]
(e) [Format specifications]
(f) [Priority: Medium (with no mention of deadline or Timeline requirements)]

**[NEXT CLEAN DD AREA NAME]**

{start_request_num + 2}. (a) [First requirement for next area]
[Continue pattern...]

GENERATE ALL {len(dd_chunk)} DD AREAS WITH PERFECT FORMATTING NOW:"""
        
        result = self._direct_llm_call(prompt, max_tokens=6000)
        return result.get("analysis", "") if result["status"] == "success" else ""
    
    def _create_clean_section_headers(self, dd_sections: List[Dict[str, Any]]) -> str:
        """Create clean section headers for all DD sections"""
        headers = []
        for section in dd_sections:
            section_name = section.get('name', 'DD Area')
            # Clean up the section name - remove any priority or formatting artifacts
            clean_name = section_name.replace('(Priority:', '').replace('High)', '').replace('Medium)', '').replace('Low)', '').strip()
            clean_name = clean_name.upper()
            headers.append(f"**{clean_name}**")
        return '\n'.join(headers)
    
    def _create_section_headers_for_chunk(self, dd_chunk: List[Dict[str, Any]]) -> str:
        """Create clean, properly formatted section headers for DD chunk"""
        headers = []
        for section in dd_chunk:
            section_name = section.get('name', 'DD Area')
            # Clean up the section name - remove any priority or formatting artifacts
            clean_name = section_name.replace('(Priority:', '').replace('High)', '').replace('Medium)', '').replace('Low)', '').strip()
            clean_name = clean_name.upper()
            headers.append(f"**{clean_name}**")
        return '\n'.join(headers)
    
    def _count_requests_in_content(self, content: str) -> int:
        """Count the number of requests in generated content"""
        import re
        # Count numbered requests (1., 2., 3., etc.)
        request_matches = re.findall(r'^\d+\.', content, re.MULTILINE)
        return len(request_matches)
    
    def _analyze_sow_scope(self, sow_content: str, dd_sections: List[Dict[str, Any]]) -> Dict[str, Any]:
        """Analyze SOW to understand actual scope and requirements"""
        
        # Count DD sections and analyze depth
        total_areas = len(dd_sections)
        
        # Estimate requests based on DD sections (3-8 requests per section typically)
        estimated_requests = total_areas * 5  # Average 5 requests per DD area
        
        # Analyze SOW content for specific focus areas
        focus_areas = []
        for section in dd_sections:
            section_name = section.get('name', '')
            if any(keyword in section_name.lower() for keyword in ['revenue', 'earnings', 'quality']):
                focus_areas.append({'name': section_name, 'priority': 'High', 'estimated_requests': 6})
            elif any(keyword in section_name.lower() for keyword in ['working capital', 'cash flow', 'balance sheet']):
                focus_areas.append({'name': section_name, 'priority': 'High', 'estimated_requests': 5})
            else:
                focus_areas.append({'name': section_name, 'priority': 'Medium', 'estimated_requests': 4})
        
        return {
            'total_areas': total_areas,
            'estimated_requests': estimated_requests,
            'focus_areas': focus_areas,
            'scope_type': 'comprehensive' if total_areas > 10 else 'focused'
        }
    
    def _generate_dynamic_irl_from_sow(self, company_name: str, financial_periods: Dict[str, str], 
                                      dd_sections: List[Dict[str, Any]], full_sow_content: str,
                                      scope_analysis: Dict[str, Any]) -> str:
        """Generate IRL dynamically based on actual SOW content"""
        
        # Format periods for prompt
        current_year = financial_periods.get('current_year', '2024-2025')
        previous_year = financial_periods.get('previous_year', '2023-2024')
        periods_text = f"FY{current_year}, FY{previous_year}"
        
        # Format the actual DD sections from SOW
        formatted_dd_scope = self._format_dd_sections_for_scope_prompt(dd_sections)
        
        # Create scope-specific prompt
        scope_instruction = self._create_scope_specific_instruction(dd_sections, scope_analysis)
        
        # Create clean section headers
        section_headers = self._create_clean_section_headers(dd_sections)
        
        # Create comprehensive prompt based on actual SOW DD areas
        prompt = f"""GENERATE INFORMATION REQUIREMENTS LIST WITH SECTION HEADERS

CRITICAL: You MUST include section headers for each DD area. Do not generate a flat list.

COMPANY: {company_name}
HISTORICAL PERIODS: {periods_text}
DD AREAS: {len(dd_sections)} specific areas

FORMAT EXAMPLE - YOU MUST FOLLOW THIS STRUCTURE:

**QUALITY OF EARNINGS ANALYSIS**

1. (a) [Specific data request for this area]
(b) [Supporting documentation needed]
(c) [Analysis methodology required]
(d) [Validation procedures]
(e) [Format specifications]
(f) [Priority: High (with no mention of deadline or Timeline requirements)]

2. (a) [Another request for same area]
(b) [Supporting documentation]
(c) [Analysis methodology]
(d) [Validation procedures]
(e) [Format specifications]
(f) [Priority: Medium (with no mention of deadline or Timeline requirements)]

**WORKING CAPITAL MANAGEMENT**

3. (a) [Request specific to working capital]
(b) [Supporting documentation]
(c) [Analysis methodology]
(d) [Validation procedures]
(e) [Format specifications]
(f) [Priority: Low (with no mention of deadline or Timeline requirements)]

MANDATORY REQUIREMENTS:
1. START each DD area with section header: **AREA NAME**
2. Generate 2-3 detailed requests per DD area
3. Number requests sequentially (1, 2, 3, etc.)
4. Include all sub-points (a) through (f)
5. Priority ONLY at end of (f): "(Priority: High/Medium/Low)"
6. NO priority in section headers

DD AREAS TO GENERATE:
{formatted_dd_scope}

Generate complete IRL with section headers for ALL {len(dd_sections)} areas listed above.
(f) [Priority: Medium (with no mention of deadline or Timeline requirements)]

**CUSTOMER ANALYSIS**

3. (a) [First request for next area]
[Continue sequential numbering...]

GENERATE COMPLETE IRL FOR ALL {len(dd_sections)} DD AREAS WITH PERFECT FORMATTING:"""
        
        # Calculate required tokens based on scope size
        required_tokens = self._calculate_required_tokens(dd_sections, scope_analysis)
        logger.info(f"📊 Using {required_tokens} tokens for {len(dd_sections)} DD areas")
        
        result = self._direct_llm_call(prompt, max_tokens=required_tokens)
        return result.get("analysis", "") if result["status"] == "success" else ""
    
    def _generate_section_b_old(self, company_name: str, financial_periods: Dict[str, str], dd_sections: List[Dict[str, Any]]) -> str:
        """Generate Section B: Profit and loss analysis (requests 16-25)"""
        
        # Extract relevant DD sections for Section B
        section_b_keywords = ["profit", "loss", "revenue", "cost", "expense", "employee", "working capital"]
        relevant_sections = [s for s in dd_sections if any(keyword in s.get("name", "").lower() for keyword in section_b_keywords)]
        
        # Format periods for prompt
        periods_text = f"FY{financial_periods.get('current_year', '2014-2015')}, FY{financial_periods.get('previous_year', '2013-2014')}"
        
        prompt = f"""GENERATE SECTION B ONLY - PROFIT AND LOSS ANALYSIS

**SECTION B: Profit and loss**

Company: {company_name}
Historical periods: {periods_text}

RELEVANT DD FOCUS AREAS:
{self._format_dd_sections_for_prompt(relevant_sections)}

Generate requests 16-25 ONLY. Continue from where Section A ended.

MANDATORY FORMAT:
16. (a) Details along with backup and transaction master of monthly revenue, ARR/MRR, marketing spends, take rate %, gross margins and KPIs in the historical period by customer and type in Excel format.
(b) By customer and type (advertiser, agencies or publishers) with detailed customer profiling and contract analysis.
(c) By billing type (minimum fee or % of take rate) with revenue mix analysis and pricing strategy documentation.
(d) By customer geo and industry verticals with market penetration analysis and competitive positioning.
(e) Type of revenue (fixed vs variable) with predictability analysis and revenue quality assessment.
(f) Revenue recognition policy and period-end cut-off procedures with supporting documentation.

17. (a) Provide copies of all contracts/agreements with top 20 customers covering 80% of revenues for the historical period.
(b) Standard terms and conditions, pricing mechanisms, payment terms, and credit periods offered to customers.
(c) Details of any volume discounts, rebates, incentives, or special pricing arrangements with calculation methodologies.
(d) Customer satisfaction scores, complaints/disputes log, and resolution tracking with impact analysis.
(e) Details of any long-term contracts, take-or-pay arrangements, or minimum purchase commitments.
(f) Customer credit assessment procedures, credit limits, and collection performance metrics.

Continue this EXACT format for requests 18-25. Focus on employee costs, operating expenses specific to {company_name}'s business model.

GENERATE ALL 10 REQUESTS (16-25) NOW - DO NOT STOP OR ASK FOR CONFIRMATION."""

        result = self._direct_llm_call(prompt)
        return result.get("analysis", "") if result["status"] == "success" else ""
    
    def _generate_section_c(self, company_name: str, financial_periods: Dict[str, str], dd_sections: List[Dict[str, Any]]) -> str:
        """Generate Section C: Balance sheet analysis (requests 26-40)"""
        
        # Extract relevant DD sections for Section C
        section_c_keywords = ["balance sheet", "assets", "liabilities", "cash", "receivables", "inventory", "debt"]
        relevant_sections = [s for s in dd_sections if any(keyword in s.get("name", "").lower() for keyword in section_c_keywords)]
        
        # Format periods for prompt
        periods_text = f"FY{financial_periods.get('current_year', '2014-2015')}, FY{financial_periods.get('previous_year', '2013-2014')}"
        
        prompt = f"""GENERATE SECTION C ONLY - BALANCE SHEET ANALYSIS

**SECTION C: Balance sheet analysis**

Company: {company_name}
Historical periods: {periods_text}

RELEVANT DD FOCUS AREAS:
{self._format_dd_sections_for_prompt(relevant_sections)}

Generate requests 26-40 ONLY (FINAL 15 REQUESTS). Continue from where Section B ended.

MANDATORY FORMAT:
26. (a) Customer-wise breakdown of accounts receivable along with agreed credit terms as at the end of the historical periods (0-30 days, 31-60 days, 61-120 days, 120-180, 180-365, greater than 365 days) in Excel format with original currency billing and exchange rates.
(b) Customer level breakdown of unbilled receivable along with ageing and subsequent invoicing details at historical balance sheet dates.
(c) Provide mapping of subsequent collections of outstanding receivables including unbilled receivables at latest balance sheet date traced to bank statements.
(d) Details of any disputed/doubtful receivables with provision calculations and recovery prospects.
(e) Credit control procedures, collection efforts tracking, and recovery performance metrics by customer category.
(f) Details of any factoring, bill discounting, or receivables financing arrangements with terms and conditions.

27. (a) Detailed fixed asset register in Excel showing asset description, location, cost, accumulated depreciation, and written down value for each asset.
(b) Asset-wise additions, deletions, and transfers during the historical period with supporting documents and approvals.
(c) Depreciation policy, useful lives, and methods adopted for each asset category with technical justification.
(d) Physical verification reports, insurance coverage details, asset tagging status, and condition assessment.
(e) Reconciliation between fixed asset register and general ledger balances with explanations for differences.
(f) Impairment assessment procedures and any impairment losses recognized during the historical period.

Continue this EXACT format for requests 28-40. Cover working capital, cash, debt, other assets, liabilities, and contingencies specific to {company_name}.

GENERATE ALL FINAL 15 REQUESTS (26-40) NOW - COMPLETE THE ENTIRE IRL."""

        result = self._direct_llm_call(prompt)
        return result.get("analysis", "") if result["status"] == "success" else ""
    
    def _generate_basic_irl_from_dd_sections(self, dd_sections: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
        """Generate basic IRL requests from DD sections when LLM parsing fails"""
        requests = []
        request_id = 1
        
        # Section A: Financial statements
        requests.append({
            "id": "A",
            "info_request": "Financial statements, MIS and other general information",
            "priority": "",
            "status": "",
            "zenalyst_remarks": "",
            "management_remarks": ""
        })
        
        requests.append({
            "id": "I",
            "info_request": "General financial information",
            "priority": "",
            "status": "",
            "zenalyst_remarks": "",
            "management_remarks": ""
        })
        
        # Generate basic requests from DD sections
        high_priority_sections = ["Quality of Earnings Analysis", "Income Statement Analysis", "Working Capital Management", "Cash Flow Analysis", "Balance Sheet Review"]
        
        for section in dd_sections:
            section_name = section["name"]
            priority = "High" if any(hp in section_name for hp in high_priority_sections) else "Medium"
            
            # Convert DD procedure to data request
            if "Quality of Earnings" in section_name:
                requests.append({
                    "id": str(request_id),
                    "info_request": f"Revenue analysis and supporting documentation for {section_name.lower()}:\n(a) Monthly revenue breakdown by customer\n(b) Customer contracts and terms\n(c) Revenue recognition policies",
                    "priority": priority,
                    "status": "",
                    "zenalyst_remarks": "",
                    "management_remarks": ""
                })
            elif "Income Statement" in section_name:
                requests.append({
                    "id": str(request_id),
                    "info_request": f"Detailed P&L analysis for {section_name.lower()}:\n(a) Cost breakdown by category\n(b) Employee cost analysis\n(c) Operating expense details",
                    "priority": priority,
                    "status": "",
                    "zenalyst_remarks": "",
                    "management_remarks": ""
                })
            elif "Working Capital" in section_name:
                requests.append({
                    "id": str(request_id),
                    "info_request": f"Working capital components for {section_name.lower()}:\n(a) Inventory aging analysis\n(b) Trade receivables aging\n(c) Trade payables analysis",
                    "priority": priority,
                    "status": "",
                    "zenalyst_remarks": "",
                    "management_remarks": ""
                })
            elif "Cash Flow" in section_name:
                requests.append({
                    "id": str(request_id),
                    "info_request": f"Cash flow supporting data for {section_name.lower()}:\n(a) Bank statements and reconciliations\n(b) Cash flow forecasts\n(c) Working capital movements",
                    "priority": priority,
                    "status": "",
                    "zenalyst_remarks": "",
                    "management_remarks": ""
                })
            else:
                # Generic request for other sections
                requests.append({
                    "id": str(request_id),
                    "info_request": f"Supporting documentation for {section_name.lower()} analysis",
                    "priority": "Medium",
                    "status": "",
                    "zenalyst_remarks": "",
                    "management_remarks": ""
                })
            
            request_id += 1
        
        return requests
    
    def _assign_request_priority(self, request_text: str) -> str:
        """Assign priority based on request content"""
        high_keywords = ["financial statements", "cash", "revenue", "receivables", "bank", "audit", "contracts"]
        medium_keywords = ["expense", "cost", "operational", "analysis", "breakdown"]
        
        request_lower = request_text.lower()
        
        for keyword in high_keywords:
            if keyword in request_lower:
                return "High"
        
        for keyword in medium_keywords:
            if keyword in request_lower:
                return "Medium"
        
        return "Low"
    
    def create_txt_output(self, irl_data: Dict[str, Any]) -> str:
        """Create TXT file for comparison with sample before Excel conversion"""
        logger.info("Creating TXT IRL output for comparison...")
        
        company_name = irl_data["company_name"]
        clean_name = company_name.replace(" ", "_").replace(".", "").replace(",", "")
        
        # Get version number
        if clean_name not in self.version_tracker:
            self.version_tracker[clean_name] = 0
        
        self.version_tracker[clean_name] += 1
        version = self.version_tracker[clean_name]
        
        # Save updated version tracker
        self._save_version_tracker()
        
        # Create filename
        filename = f"{clean_name}_IRL_v{version}.txt"
        
        # Create TXT content
        txt_content = []
        txt_content.append("INFORMATION REQUIREMENTS LIST")
        txt_content.append("=" * 50)
        txt_content.append(f"Company: {company_name}")
        txt_content.append(f"Generated: {datetime.now().strftime('%B %d, %Y')}")
        txt_content.append("")
        txt_content.append("Historical period: FY2014-2015, FY2013-2014")
        txt_content.append("Balance sheet date: March 31, 2015")
        txt_content.append("Information on consolidated basis wherever applicable")
        txt_content.append("")
        txt_content.append("=" * 50)
        txt_content.append("")
        
        # Add structured data
        structured_data = irl_data["irl_data"]
        for item in structured_data:
            item_id = item.get("id", "")
            info_request = item.get("info_request", "")
            priority = item.get("priority", "")
            
            # Handle section headers (no ID) and regular requests (with ID)
            if info_request:
                if item_id:  # Regular request with number
                    if priority:
                        txt_content.append(f"{item_id}. {info_request} (Priority: {priority})")
                    else:
                        txt_content.append(f"{item_id}. {info_request}")
                else:  # Section header or other info
                    txt_content.append(info_request)
                txt_content.append("")
        
        # Save file
        with open(filename, 'w', encoding='utf-8') as f:
            f.write('\n'.join(txt_content))
            
        logger.info(f"📄 IRL saved as: {filename}")
        
        return filename
    
    def create_excel_output(self, irl_data: Dict[str, Any]) -> str:
        """Create Excel file from IRL data maintaining proper structure"""
        logger.info("Creating Excel IRL output...")
        
        company_name = irl_data["company_name"]
        clean_name = company_name.replace(" ", "_").replace(".", "").replace(",", "")
        
        # Use the same version as TXT file
        version = self.version_tracker.get(clean_name, 1)
        
        # Create Excel filename
        excel_filename = f"{clean_name}_IRL_v{version}.xlsx"
        
        # Create a new workbook
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        from openpyxl.utils.dataframe import dataframe_to_rows
        import pandas as pd
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Information Requirements List"
        
        # Set up styles
        header_font = Font(bold=True, size=14, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        section_font = Font(bold=True, size=12, color="FFFFFF")
        section_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
        normal_font = Font(size=11)
        border = Border(left=Side(style='thin'), right=Side(style='thin'), 
                       top=Side(style='thin'), bottom=Side(style='thin'))
        
        # Header section
        ws.merge_cells('A1:E1')
        ws['A1'] = "INFORMATION REQUIREMENTS LIST"
        ws['A1'].font = header_font
        ws['A1'].fill = header_fill
        ws['A1'].alignment = Alignment(horizontal='center')
        
        ws.merge_cells('A2:E2')
        ws['A2'] = f"Company: {company_name}"
        ws['A2'].font = Font(bold=True, size=12)
        ws['A2'].alignment = Alignment(horizontal='center')
        
        ws.merge_cells('A3:E3')
        ws['A3'] = f"Generated: {datetime.now().strftime('%B %d, %Y')}"
        ws['A3'].font = normal_font
        ws['A3'].alignment = Alignment(horizontal='center')
        
        # FIXED: Add periods info properly from financial_periods
        financial_periods = irl_data.get("financial_periods", {})
        current_year = financial_periods.get("current_year", "N/A")
        previous_year = financial_periods.get("previous_year", "N/A")
        balance_date = financial_periods.get("balance_sheet_date", "N/A")
        
        ws.merge_cells('A5:E5')
        ws['A5'] = f"Historical period: {previous_year}, {current_year}"
        ws['A5'].font = normal_font
        
        ws.merge_cells('A6:E6')
        ws['A6'] = f"Balance sheet date: {balance_date}"
        ws['A6'].font = normal_font
        
        ws.merge_cells('A7:E7')
        ws['A7'] = "Information on consolidated basis wherever applicable"
        ws['A7'].font = normal_font
        
        # Column headers
        row = 9  # FIXED: Updated row position to account for additional header rows
        headers = ['S.No.', 'Section', 'Information Requirement', 'Priority', 'Comments']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.font = section_font
            cell.fill = section_fill
            cell.alignment = Alignment(horizontal='center')
            cell.border = border
        
        # Set column widths
        ws.column_dimensions['A'].width = 8
        ws.column_dimensions['B'].width = 25
        ws.column_dimensions['C'].width = 80
        ws.column_dimensions['D'].width = 12
        ws.column_dimensions['E'].width = 20
        
        # COMPLETELY REWRITTEN: Process IRL data properly
        row = 10
        item_counter = 1
        structured_data = irl_data["irl_data"]
        
        # Group data by sections and combine sub-points
        current_section = ""
        section_items = {}  # Track items per section
        
        for item in structured_data:
            info_request = item.get("info_request", "")
            priority = item.get("priority", "")
            
            if not info_request:
                continue
            
            # Skip historical period items that were incorrectly added to data
            if any(skip_text in info_request.lower() for skip_text in ['historical period:', 'balance sheet date:', 'information on consolidated basis']):
                continue
            
            # Check if this is a section header (starts with **)
            if info_request.startswith("**") and info_request.endswith("**"):
                current_section = info_request.replace("**", "").strip()
                section_items[current_section] = []
                continue
            
            # Regular item - collect all sub-points for this section
            if current_section:
                lines = info_request.split('\n')
                main_req = lines[0] if lines else info_request
                
                # Clean up numbering if present
                if main_req.strip() and main_req.strip()[0].isdigit():
                    main_req = '. '.join(main_req.split('. ')[1:]) if '. ' in main_req else main_req
                
                # Collect all sub-points (a), (b), (c), etc.
                sub_points = []
                if len(lines) > 1:
                    for line in lines[1:]:
                        line = line.strip()
                        if line and (line.startswith('(') or any(line.startswith(f'({letter})') for letter in 'abcdefghijklmnop')):
                            sub_points.append(line)
                
                if current_section not in section_items:
                    section_items[current_section] = []
                
                section_items[current_section].extend(sub_points)
        
        # Now write the properly grouped data
        for section_name, sub_points in section_items.items():
            if not sub_points:
                continue
                
            # Combine all sub-points for this section into one cell
            combined_requirements = '\n'.join(sub_points)
            
            # Add row for this section
            ws.cell(row=row, column=1, value=item_counter).font = normal_font
            ws.cell(row=row, column=1).border = border
            ws.cell(row=row, column=1).alignment = Alignment(horizontal='center')
            
            ws.cell(row=row, column=2, value=section_name).font = normal_font
            ws.cell(row=row, column=2).border = border
            
            # FIXED: Put combined requirements in Information Requirement column
            ws.cell(row=row, column=3, value=combined_requirements).font = normal_font
            ws.cell(row=row, column=3).border = border
            ws.cell(row=row, column=3).alignment = Alignment(wrap_text=True, vertical='top')
            
            # Assign priority based on section name
            if any(keyword in section_name.upper() for keyword in ['REVENUE', 'EARNINGS', 'CASH', 'BALANCE']):
                priority = "High"
            elif any(keyword in section_name.upper() for keyword in ['WORKING CAPITAL', 'INCOME']):
                priority = "Medium"
            else:
                priority = "Low"
                
            ws.cell(row=row, column=4, value=priority).font = normal_font
            ws.cell(row=row, column=4).border = border
            ws.cell(row=row, column=4).alignment = Alignment(horizontal='center')
            
            # FIXED: Leave comments column empty as requested
            ws.cell(row=row, column=5, value="").border = border
            
            # Set row height based on content
            estimated_lines = len(sub_points)
            ws.row_dimensions[row].height = max(20 * estimated_lines, 30)
            
            row += 1
            item_counter += 1
        
        # Save Excel file
        try:
            wb.save(excel_filename)
            logger.info(f"📊 Excel IRL saved as: {excel_filename}")
            return excel_filename
        except Exception as e:
            logger.error(f"Failed to save Excel file: {e}")
            return None
    
    def _get_default_irl_prompt(self) -> str:
        """Default IRL generation prompt if file not found"""
        return """
You are an expert financial due diligence specialist. Convert the provided due diligence scope into specific, actionable data requests.

TASK: Generate Information Requirements List (IRL) with specific document and data requests.

STRUCTURE REQUIREMENTS:
- Section A: Financial statements, MIS and general information
- Section B: Profit and loss analysis  
- Section C: Balance sheet analysis

For each DD procedure, create specific data requests that include:
- Exact documents needed
- Time periods required
- Data formats (Excel preferred)
- Specific breakdowns needed

PRIORITY ASSIGNMENT:
- High: Core financials, cash, revenue, receivables
- Medium: Supporting operational data
- Low: Nice-to-have information

Generate practical, implementable data requests that will enable the DD analysis described in the scope.
"""
    
    def process_sow_to_irl(self, sow_file_path: str) -> Dict[str, Any]:
        """Complete pipeline: SOW file → IRL Excel"""
        logger.info(f"🚀 Starting IRL generation from SOW: {sow_file_path}")
        
        # Step 1: Read SOW output
        sow_data = self.read_sow_output(sow_file_path)
        if sow_data["status"] != "success":
            return sow_data
        
        logger.info(f"✅ SOW parsed - Company: {sow_data['company_name']}")
        
        # Step 2: Generate IRL
        irl_result = self.generate_irl_from_sow(sow_data)
        if irl_result["status"] != "success":
            return irl_result
        
        logger.info("✅ IRL content generated")
        
        # Step 3: Create both TXT and Excel outputs
        logger.info("Creating TXT IRL output for comparison...")
        txt_filename = self.create_txt_output(irl_result)
        
        logger.info("Creating Excel IRL output...")
        excel_filename = self.create_excel_output(irl_result)
        
        return {
            "status": "success",
            "company_name": sow_data["company_name"],
            "sow_file": sow_file_path,
            "irl_txt_file": txt_filename,
            "irl_excel_file": excel_filename,
            "financial_periods": sow_data["financial_periods"],
            "dd_sections_count": len(sow_data["dd_sections"])
        }

def main():
    """Command line interface for IRL generation"""
    if len(sys.argv) < 2:
        print("📋 IRL DD PIPELINE")
        print("=" * 50)
        print("Usage: python irl_dd_pipeline.py <sow_output_file>")
        print("\nExample:")
        print('  python irl_dd_pipeline.py "../SOW LLM/COMPANY_DD_SOW_v1.txt"')
        print("\nFeatures:")
        print("  ✅ Reads SOW LLM output automatically")
        print("  ✅ Extracts company name and financial periods")
        print("  ✅ Generates specific data requests")
        print("  ✅ Creates Excel IRL with version tracking")
        return
    
    sow_file_path = sys.argv[1]
    
    if not os.path.exists(sow_file_path):
        print(f"❌ Error: SOW file not found: {sow_file_path}")
        return
    
    # Initialize pipeline
    pipeline = IRLDueDiligencePipeline()
    
    print("🚀 IRL DD PIPELINE")
    print("=" * 50)
    print(f"📄 Processing SOW: {sow_file_path}")
    
    # Process SOW to IRL
    result = pipeline.process_sow_to_irl(sow_file_path)
    
    if result["status"] == "success":
        print("\n✅ IRL GENERATION COMPLETED!")
        print("=" * 50)
        print(f"🏢 Company: {result['company_name']}")
        print(f"📊 SOW Source: {result['sow_file']}")
        print(f"📋 IRL TXT Output: {result['irl_txt_file']}")
        print(f"📊 IRL Excel Output: {result['irl_excel_file']}")
        print(f"📅 Periods: {result['financial_periods']}")
        print(f"🔍 DD Sections: {result['dd_sections_count']}")
    else:
        print(f"\n❌ IRL GENERATION FAILED")
        print("=" * 30)
        print(f"Error: {result['error']}")

if __name__ == "__main__":
    main()